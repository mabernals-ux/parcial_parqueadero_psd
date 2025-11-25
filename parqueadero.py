from flask import Flask, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import re
import openpyxl
from io import BytesIO
from sqlalchemy.exc import IntegrityError, DataError, OperationalError
from werkzeug.exceptions import MethodNotAllowed
import math
from flask_cors import CORS

# Variables globales para RFID
ultimo_uid = None
ultimo_tipo = None
rfid_timestamp = None


# ======================================================
# CONFIGURACIÓN APP Y BASE DE DATOS
# ======================================================

app = Flask(__name__)
CORS(app)  # permitir llamados desde cualquier origen

app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:8559@localhost:5432/parqueadero'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ======================================================
# MODELOS
# ======================================================

class TipoDocumento(db.Model):
    __tablename__ = 'tipos_documento'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False)
    usuarios = db.relationship('Usuario', backref='tipo_documento', lazy=True)

class TipoVehiculo(db.Model):
    __tablename__ = 'tipos_vehiculo'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(20), unique=True, nullable=False)
    vehiculos = db.relationship('Vehiculo', backref='tipo_vehiculo_ref', lazy=True)

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    tipo_documento_id = db.Column(db.Integer, db.ForeignKey('tipos_documento.id'), nullable=False)
    numero_identificacion = db.Column(db.String(20), unique=True, nullable=False)
    saldo = db.Column(db.Float, default=0.0)

class Vehiculo(db.Model):
    __tablename__ = 'vehiculos'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    placa = db.Column(db.String(20), unique=True, nullable=False)
    tipo_vehiculo_id = db.Column(db.Integer, db.ForeignKey('tipos_vehiculo.id'), nullable=False)
    uid_rfid = db.Column(db.String(20), unique=True)
    usuario = db.relationship("Usuario", backref="vehiculos", lazy=True)

class Espacio(db.Model):
    __tablename__ = 'espacios'
    id = db.Column(db.Integer, primary_key=True)
    tipo_vehiculo_id = db.Column(db.Integer, db.ForeignKey('tipos_vehiculo.id'), nullable=False)
    estado = db.Column(db.Boolean, default=False)  # False = libre, True = ocupado
    vehiculo_id = db.Column(db.Integer, db.ForeignKey('vehiculos.id'), nullable=True)

class Registro(db.Model):
    __tablename__ = 'registros'
    id = db.Column(db.Integer, primary_key=True)
    vehiculo_id = db.Column(db.Integer, db.ForeignKey('vehiculos.id'), nullable=False)
    espacio_id = db.Column(db.Integer, db.ForeignKey('espacios.id'), nullable=False)
    hora_ingreso = db.Column(db.DateTime, nullable=False)
    hora_salida = db.Column(db.DateTime, nullable=True)
    tiempo_duracion = db.Column(db.Float, nullable=True)  # en horas
    total_pago = db.Column(db.Float, nullable=True)
    vehiculo = db.relationship("Vehiculo", backref="registros", lazy=True)
    espacio = db.relationship("Espacio", backref="registros", lazy=True)

class Tarifa(db.Model):
    __tablename__ = 'tarifas'
    id = db.Column(db.Integer, primary_key=True)
    tipo_vehiculo_id = db.Column(db.Integer, db.ForeignKey('tipos_vehiculo.id'), nullable=False)
    tarifa_hora = db.Column(db.Float, nullable=False)

class Recarga(db.Model):
    __tablename__ = 'recargas'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    saldo_anterior = db.Column(db.Float, nullable=True)
    monto_recargado = db.Column(db.Float, nullable=False)
    saldo_final = db.Column(db.Float, nullable=True)
    referencia = db.Column(db.String(50), nullable=True)
    fecha_recarga = db.Column(db.DateTime, default=datetime.now)
    usuario = db.relationship('Usuario', backref='recargas', lazy=True)

class ValorMinimo(db.Model):
    __tablename__ = 'valor_minimo'
    id = db.Column(db.Integer, primary_key=True)
    valor = db.Column(db.Float, unique=True, nullable=False)
    def __repr__(self):
        return f"<ValorMinimo {self.valor}>"

# Crear tablas
with app.app_context():
    db.create_all()

# ======================================================
# EXCEPCIONES PERSONALIZADAS
# ======================================================
class VehiculoNoRegistradoError(Exception): pass
class SaldoInsuficienteError(Exception): pass
class EspacioNoDisponibleError(Exception): pass

# ======================================================
# FUNCIONES AUXILIARES
# ======================================================
def calcular_tarifa(vehiculo_id, minutos):
    vehiculo = Vehiculo.query.get(vehiculo_id)
    if not vehiculo:
        return 0
    tarifa = Tarifa.query.filter_by(tipo_vehiculo_id=vehiculo.tipo_vehiculo_id).first()
    if not tarifa:
        return 0
    tarifa_por_minuto = tarifa.tarifa_hora
    total = round(tarifa_por_minuto * minutos, 2)
    return total

# ======================================================
# ENDPOINTS (Usuarios, Vehículos, Parqueadero)
# ======================================================
# --------------------------
# ENDPOINTS
# --------------------------

# Registrar usuario
@app.route('/usuarios', methods=['POST'])
def registrar_usuario():
    try:
        data = request.get_json()
        nombre = data.get("nombre")
        tipo_doc = data.get("tipo_documento")   # CC, TI, PAS, NIT
        numero_id = data.get("numero_identificacion")
        saldo = data.get("saldo", 0.0)

        if not nombre or not tipo_doc or not numero_id:
            return jsonify({"message": "Faltan datos requeridos"}), 400

        # -------------------------------
        # VALIDAR SALDO INICIAL CON VALOR MINIMO DESDE BD
        # -------------------------------
        try:
            saldo = float(saldo)
        except:
            return jsonify({"message": "El saldo debe ser un número válido"}), 400

        # Obtener valor mínimo desde la tabla
        valor_minimo_obj = ValorMinimo.query.first()
        valor_minimo = valor_minimo_obj.valor if valor_minimo_obj else 5000  # fallback

        if saldo < valor_minimo:
            return jsonify({"message": f"El saldo inicial no puede ser menor a {valor_minimo}"}), 400

        # Validar nombre solo letras y espacios
        if not re.match(r'^[A-Za-zÁÉÍÓÚáéíóúÑñ\s]+$', nombre):
            return jsonify({"message": "El nombre solo puede contener letras y espacios"}), 400

        # Buscar el tipo de documento en la BD
        tipo_doc = tipo_doc.upper().strip()
        tipo_doc_obj = TipoDocumento.query.filter_by(nombre=tipo_doc).first()
        if not tipo_doc_obj:
            return jsonify({"message": "Tipo de documento no válido. Use CC, PAS, TI o NIT"}), 400

        # -------------------------------
        # VALIDACIONES SEGÚN EL DOCUMENTO
        # -------------------------------
        if tipo_doc == "CC":  # Cédula 8-10 números
            if not re.fullmatch(r'^\d{8,10}$', numero_id):
                return jsonify({"message": "La cédula debe tener entre 8 y 10 números"}), 400

        elif tipo_doc == "PAS":  # Pasaporte 2 letras + 6 números
            if not re.fullmatch(r'^[A-Za-z]{2}\d{6}$', numero_id):
                return jsonify({"message": "El pasaporte debe tener 2 letras seguidas de 6 números"}), 400

        elif tipo_doc == "TI":  # Tarjeta de Identidad 10 números
            if not re.fullmatch(r'^\d{10}$', numero_id):
                return jsonify({"message": "La tarjeta de identidad debe tener 10 números"}), 400

        elif tipo_doc == "NIT":  # NIT 9 números + guion + 1 número
            if not re.fullmatch(r'^\d{9}-\d{1}$', numero_id):
                return jsonify({"message": "El NIT debe tener 9 números, guion y 1 número (ej: 123456789-0)"}), 400

        # -------------------------------
        # VALIDAR DUPLICADOS
        # -------------------------------
        if Usuario.query.filter_by(numero_identificacion=numero_id).first():
            return jsonify({"message": "Número de identificación ya existe"}), 400

        # -------------------------------
        # CREAR USUARIO (guardando el ID real del tipo_doc)
        # -------------------------------
        nuevo = Usuario(
            nombre=nombre.strip(),
            tipo_documento_id=tipo_doc_obj.id,  # Guardamos el id (1,2,3,4)
            numero_identificacion=numero_id,
            saldo=saldo
        )
        db.session.add(nuevo)
        db.session.commit()

        return jsonify({
            "message": f"Usuario {nombre.strip()} registrado exitosamente",
            "usuario_id": nuevo.id
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"Error inesperado: {str(e)}"}), 500


# Registrar vehículo
@app.route('/vehiculos', methods=['POST'])
def registrar_vehiculo():
    global ultimo_uid  # Usar la variable global que contiene el UID leído
    try:
        data = request.get_json()
        placa = data.get("placa")
        tipo_vehiculo_id = data.get("tipo_vehiculo_id")
        numero_identificacion = data.get("numero_identificacion")  

        # Tomar UID del último RFID leído
        uid_rfid = ultimo_uid

        # Verificar si ya se leyó un UID
        if not uid_rfid:
            return jsonify({"message": "No se ha leído ningún RFID, lea la tarjeta primero"}), 400

        # Verificar si el UID ya está registrado
        if Vehiculo.query.filter_by(uid_rfid=uid_rfid).first():
            return jsonify({"message": "Este RFID ya está asignado a otro vehículo"}), 400

        # Validar datos obligatorios
        if not placa or not tipo_vehiculo_id or not numero_identificacion:
            return jsonify({"message": "Faltan datos requeridos"}), 400

        # Validar usuario por documento
        usuario = Usuario.query.filter_by(numero_identificacion=numero_identificacion).first()
        if not usuario:
            return jsonify({"message": "Usuario no encontrado con ese documento"}), 404

        # Validar tipo de vehículo
        tipo_vehiculo = TipoVehiculo.query.get(tipo_vehiculo_id)
        if not tipo_vehiculo:
            return jsonify({"message": "Tipo de vehículo no válido"}), 400

        # Validar formato de placa
        if tipo_vehiculo_id == 1:  # Carro
            if not re.fullmatch(r'^[A-Z]{3}[0-9]{3}$', placa):
                return jsonify({"message": "Placa inválida. Formato carro: ABC123"}), 400
        elif tipo_vehiculo_id == 2:  # Moto
            if not re.fullmatch(r'^[A-Z]{3}[0-9]{2}[A-Z]$', placa):
                return jsonify({"message": "Placa inválida. Formato moto: ABC12D"}), 400
        else:
            return jsonify({"message": "Tipo de vehículo no soportado"}), 400

        # Validar placa única
        if Vehiculo.query.filter_by(placa=placa).first():
            return jsonify({"message": "La placa ya está registrada"}), 400

        # Registrar vehículo
        nuevo = Vehiculo(
            placa=placa,
            tipo_vehiculo_id=tipo_vehiculo_id,
            usuario_id=usuario.id,
            uid_rfid=uid_rfid
        )
        db.session.add(nuevo)
        db.session.commit()

        # Limpiar UID global para no reutilizarlo
        ultimo_uid = None

        return jsonify({
            "message": f"Vehículo {placa} registrado exitosamente",
            "usuario": usuario.nombre,
            "documento": usuario.numero_identificacion,
            "uid_rfid": uid_rfid
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"Error inesperado: {str(e)}"}), 500


# ======================================================
# ENDPOINT PARA LEER RFID TEMPORAL
# ======================================================
# Al inicio de tu archivo
ultimo_uid = None  # inicializar variable global

# ====================================================== 
# ENDPOINT PARA LEER RFID TEMPORAL
# ======================================================


@app.route("/rfid/ultimo", methods=["GET"])
def rfid_ultimo():
    global ultimo_uid, ultimo_tipo, rfid_timestamp
    if ultimo_uid is None:
        return jsonify({
            "uid": None,
            "tipo": None,
            "timestamp": None,
            "message": "No se ha leído ningún RFID todavía"
        }), 200
    return jsonify({
        "uid": ultimo_uid,
        "tipo": ultimo_tipo,
        "timestamp": rfid_timestamp.strftime("%Y-%m-%d %H:%M:%S") if rfid_timestamp else None
    }), 200

# -------------------------------------------
# LISTAR VEHÍCULOS
# -------------------------------------------
@app.route('/vehiculos', methods=['GET'])
def obtener_vehiculos():
    try:
        vehiculos = Vehiculo.query.all()
        formato = request.args.get("formato")

        if formato and formato.lower() == "excel":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vehículos"

            # Encabezados
            ws.append(["ID", "Placa", "Tipo Vehículo", "Propietario", "RFID"])

            # Filas con datos
            for v in vehiculos:
                ws.append([
                    v.id,
                    v.placa,
                    v.tipo_vehiculo_ref.nombre,
                    v.usuario.nombre,
                    getattr(v, "uid_rfid", "")  # 👈 muestra la columna uid.rfid
                ])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                download_name="vehiculos.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # JSON normal
        vehiculos_list = []
        for v in vehiculos:
            vehiculos_list.append({
                "id": v.id,
                "placa": v.placa,
                "tipo_vehiculo": v.tipo_vehiculo_ref.nombre,
                "propietario": v.usuario.nombre,
                "rfid": getattr(v, "uid_rfid", "")  # 👈 incluirlo también aquí
            })
        return jsonify(vehiculos_list), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500


#Asignar espacio
@app.route('/parqueadero/asignar', methods=['POST'])
def asignar():
    data = request.get_json()
    placa = data.get("placa")

    if not placa:
        return {"message": "Debe enviar la placa del vehículo"}, 400

    return asignar_espacio(placa)


def asignar_espacio(placa):
    # Buscar vehículo
    vehiculo = Vehiculo.query.filter_by(placa=placa).first()
    if not vehiculo:
        return {"message": f"Vehículo con placa {placa} no existe"}, 404

    # Verificar si ya está en un espacio ocupado
    espacio_ocupado = Espacio.query.filter_by(vehiculo_id=vehiculo.id, estado=True).first()
    if espacio_ocupado:
        return {"message": f"El vehículo {placa} ya está en el espacio {espacio_ocupado.id}"}, 400

    # Buscar espacio libre del mismo tipo
    espacio = Espacio.query.filter_by(tipo_vehiculo_id=vehiculo.tipo_vehiculo_id, estado=False).first()
    if not espacio:
        return {"message": "No hay espacios disponibles para este tipo de vehículo"}, 400

    # Asignar espacio
    espacio.estado = True
    espacio.vehiculo_id = vehiculo.id
    db.session.commit()
    
    # Crear registro de ingreso
    hora_asignacion = datetime.now()
    nuevo_registro = Registro(
        vehiculo_id=vehiculo.id,
        espacio_id=espacio.id,
        hora_ingreso=hora_asignacion
    )
    db.session.add(nuevo_registro)
    db.session.commit()

    return {
        "message": f"Espacio {espacio.id} asignado al vehículo {placa}",
        "hora_asignacion": hora_asignacion.strftime("%Y-%m-%d %H:%M:%S")
    }, 200

# GET para mostrar el estado completo del parqueadero
@app.route('/parqueadero/estado', methods=['GET'])
def estado_parqueadero():
    try:
        espacios = Espacio.query.order_by(Espacio.id).all()
        formato = request.args.get("formato")

        if formato and formato.lower() == "excel":
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Estado Parqueadero"

            # Encabezados
            ws.append([
                "ID Espacio",
                "Ocupado",
                "Placa Vehículo"
            ])

            # Filas con datos
            for e in espacios:
                if e.vehiculo_id:
                    v = Vehiculo.query.get(e.vehiculo_id)
                    ws.append([e.id, "Sí", v.placa])
                else:
                    ws.append([e.id, "No", None])

            # Guardar en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                download_name="estado_parqueadero.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Si no piden Excel → devolver JSON
        estado = {}
        for e in espacios:
            if e.vehiculo_id:
                v = Vehiculo.query.get(e.vehiculo_id)
                estado[e.id] = v.placa
            else:
                estado[e.id] = None

        return jsonify({
            "estado_parqueadero": estado
        }), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500


# Registro salida con control de saldo

@app.route('/parqueadero/movimiento', methods=['POST'])
def movimiento():
    try:
        data = request.get_json()
        placa = data.get("placa")
        if not placa:
            return {"message": "Debe enviar la placa"}, 400

        return registrar_salida(placa)

    except VehiculoNoRegistradoError as e:
        return {"error": str(e)}, 404
    except Exception as e:
        return {"error": f"Error inesperado: {str(e)}"}, 500


def registrar_salida(placa):
    # Buscar vehículo
    vehiculo = Vehiculo.query.filter_by(placa=placa).first()
    if not vehiculo:
        raise VehiculoNoRegistradoError(f"El vehículo con placa {placa} no está registrado.")

    # Buscar registro de ingreso activo (sin salida)
    registro_activo = Registro.query.filter_by(vehiculo_id=vehiculo.id, hora_salida=None).first()
    if not registro_activo:
        return {"message": f"El vehículo {placa} no tiene un ingreso activo"}, 400

    # Calcular tiempo
    hora_salida = datetime.now()
    delta = hora_salida - registro_activo.hora_ingreso
    minutos_exactos = delta.total_seconds() / 60  # tiempo en minutos
    minutos = math.ceil(minutos_exactos)  # Redondear hacia arriba

    # Calcular el total a pagar con la tabla tarifas
    total_pago = calcular_tarifa(vehiculo.id, minutos)

    # Verificar saldo del usuario
    usuario = vehiculo.usuario
    saldo_anterior = usuario.saldo

    if saldo_anterior < total_pago:
        return {
            "message": "Saldo insuficiente, debe recargar",
            "saldo_actual": saldo_anterior,
            "total_a_pagar": total_pago,
            "faltante": round(total_pago - saldo_anterior, 2)
        }, 400

    # Registrar salida
    registro_activo.hora_salida = hora_salida
    registro_activo.tiempo_duracion = minutos
    registro_activo.total_pago = total_pago

    # Descontar saldo
    usuario.saldo -= total_pago
    saldo_final = usuario.saldo

    # Liberar espacio
    espacio = Espacio.query.get(registro_activo.espacio_id)
    espacio.estado = False
    espacio.vehiculo_id = None

    db.session.commit()

    return {
        "message": f"Vehículo {placa} salió.",
        "hora_ingreso": registro_activo.hora_ingreso.strftime("%Y-%m-%d %H:%M:%S"),
        "hora_salida": registro_activo.hora_salida.strftime("%Y-%m-%d %H:%M:%S"),
        "tiempo_total_minutos": minutos,
        "total_a_pagar": total_pago,
        "saldo_anterior": saldo_anterior,
        "saldo_final": saldo_final
    }, 200


# Consultar todos los usuarios

@app.route('/usuarios', methods=['GET'])
def obtener_usuarios():
    try:
        usuarios = Usuario.query.all()

        # Verificar si se pidió Excel
        formato = request.args.get("formato")

        if formato and formato.lower() == "excel":
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Usuarios"

            # Encabezados
            ws.append(["ID", "Nombre", "Tipo Documento", "Número Identificación", "Saldo"])

            # Filas
            for u in usuarios:
                ws.append([
                    u.id,
                    u.nombre,
                    u.tipo_documento.nombre,
                    u.numero_identificacion,
                    u.saldo
                ])

            # Guardar en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Enviar archivo
            return send_file(
                output,
                download_name="usuarios.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Si no pidieron Excel, devolver JSON
        usuarios_list = []
        for usuario in usuarios:
            usuarios_list.append({
                "id": usuario.id,
                "nombre": usuario.nombre,
                "tipo_documento": usuario.tipo_documento.nombre,
                "numero_identificacion": usuario.numero_identificacion,
                "saldo": usuario.saldo
            })
        return jsonify(usuarios_list), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500

    
# Endpoint para detalle de usuario y sus vehículos

@app.route('/usuario/<int:usuario_id>/detalle', methods=['GET'])
def obtener_detalle_usuario(usuario_id):
    usuario = Usuario.query.get(usuario_id)
    if not usuario:
        return jsonify({"message": "Usuario no encontrado"}), 404

    vehiculos_info = []
    for vehiculo in usuario.vehiculos:
        vehiculos_info.append({
            "placa": vehiculo.placa,
            "tipo_vehiculo": vehiculo.tipo_vehiculo_ref.nombre
        })

    resultado = {
        "usuario_id": usuario.id,
        "nombre": usuario.nombre,
        "tipo_documento": usuario.tipo_documento.nombre,
        "numero_identificacion": usuario.numero_identificacion,
        "saldo": usuario.saldo,
        "vehiculos": vehiculos_info
    }
    return jsonify(resultado), 200

# Consultar tarifas
@app.route('/tarifas', methods=['GET'])
def obtener_tarifas():
    try:
        tarifas = Tarifa.query.all()
        formato = request.args.get("formato")

        if formato and formato.lower() == "excel":
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tarifas"

            # Encabezados
            ws.append([
                "ID",
                "Tipo Vehículo",
                "Tarifa por Hora"
            ])

            # Filas con datos
            for tarifa in tarifas:
                tipo = TipoVehiculo.query.get(tarifa.tipo_vehiculo_id)
                ws.append([
                    tarifa.id,
                    tipo.nombre if tipo else "Desconocido",
                    tarifa.tarifa_hora
                ])

            # Guardar en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                download_name="tarifas.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Si no piden Excel → devolver JSON
        tarifas_list = []
        for tarifa in tarifas:
            tipo = TipoVehiculo.query.get(tarifa.tipo_vehiculo_id)
            tarifas_list.append({
                "id": tarifa.id,
                "tipo_vehiculo": tipo.nombre if tipo else None,
                "tarifa_hora": tarifa.tarifa_hora
            })
        return jsonify(tarifas_list), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500

    
# Consultar registros
@app.route('/registros', methods=['GET'])
def obtener_registros():
    try:
        registros = Registro.query.all()
        formato = request.args.get("formato")

        if formato and formato.lower() == "excel":
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registros"

            # Encabezados
            ws.append([
                "ID",
                "Placa",
                "Propietario",
                "Espacio",
                "Hora Ingreso",
                "Hora Salida",
                "Duración (minutos)",
                "Total Pago"
            ])

            # Filas con datos
            for r in registros:
                ws.append([
                    r.id,
                    r.vehiculo.placa,
                    r.vehiculo.usuario.nombre,
                    r.espacio_id,
                    r.hora_ingreso.strftime("%Y-%m-%d %H:%M:%S"),
                    r.hora_salida.strftime("%Y-%m-%d %H:%M:%S") if r.hora_salida else "En curso",
                    round(r.tiempo_duracion, 2) if r.tiempo_duracion else None,
                    r.total_pago if r.total_pago else 0.0
                ])

            # Guardar en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                download_name="registros.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Si no pidieron Excel → devolver JSON
        registros_list = []
        for r in registros:
            registros_list.append({
                "id": r.id,
                "placa": r.vehiculo.placa,
                "propietario": r.vehiculo.usuario.nombre,
                "espacio": r.espacio_id,
                "hora_ingreso": r.hora_ingreso.strftime("%Y-%m-%d %H:%M:%S"),
                "hora_salida": r.hora_salida.strftime("%Y-%m-%d %H:%M:%S") if r.hora_salida else None,
                "duracion_minutos": round(r.tiempo_duracion, 2) if r.tiempo_duracion else None,
                "total_pago": r.total_pago
            })
        return jsonify(registros_list), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500


#Calcular tarifa
def calcular_tarifa(vehiculo_id, minutos):
    vehiculo = Vehiculo.query.get(vehiculo_id)
    if not vehiculo:
        return 0

    tarifa = Tarifa.query.filter_by(tipo_vehiculo_id=vehiculo.tipo_vehiculo_id).first()
    if not tarifa:
        return 0

    tarifa_por_minuto = tarifa.tarifa_hora # tarifa_hora se interpreta como costo por minuto 
    total = round(tarifa_por_minuto * minutos, 2)
    return total

#generar reporte de pagos

@app.route('/reportes/pagos', methods=['GET'])
def reporte_pagos():
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        tipo_vehiculo_id = request.args.get('tipo_vehiculo_id')

        query = Registro.query.filter(Registro.hora_salida.isnot(None))

        if fecha_inicio:
            query = query.filter(Registro.hora_ingreso >= datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        if fecha_fin:
            query = query.filter(Registro.hora_ingreso <= datetime.strptime(fecha_fin, '%Y-%m-%d'))

        if tipo_vehiculo_id:
            query = query.join(Vehiculo).filter(Vehiculo.tipo_vehiculo_id == int(tipo_vehiculo_id))

        registros = query.order_by(Registro.hora_ingreso.desc()).all()

        reporte = []
        total_ingresos = 0.0
        for reg in registros:
            total_pagado = reg.total_pago or 0.0
            total_ingresos += total_pagado
            reporte.append({
                "placa": reg.vehiculo.placa,
                "tipo_vehiculo": reg.vehiculo.tipo_vehiculo_ref.nombre,
                "tiempo_minutos": reg.tiempo_duracion,
                "total_pagado": total_pagado,
                "hora_ingreso": reg.hora_ingreso.strftime("%Y-%m-%d %H:%M"),
                "hora_salida": reg.hora_salida.strftime("%Y-%m-%d %H:%M"),
                "propietario": reg.vehiculo.usuario.nombre
            })

        return jsonify({
            "registros": reporte,
            "total_registros": len(reporte),
            "total_ingresos": round(total_ingresos, 2)
        }), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500


# Recargar saldo

@app.route('/usuarios/recargar', methods=['POST'])
def recargar_saldo():
    try:
        data = request.get_json() or request.form

        numero_identificacion = data.get('numero_identificacion')
        monto = data.get('monto')

        if not numero_identificacion or not monto:
            return jsonify({"error": "Faltan datos requeridos"}), 400

        usuario = Usuario.query.filter_by(numero_identificacion=str(numero_identificacion)).first()
        if not usuario:
            return jsonify({"error": "Usuario no encontrado"}), 404

        monto = float(monto)
        saldo_anterior = usuario.saldo
        saldo_final = saldo_anterior + monto

        # ✅ Generar referencia automática única
        fecha_actual = datetime.now().strftime("%Y%m%d-%H%M%S")
        referencia = f"REC-{fecha_actual}-{usuario.numero_identificacion}"

        # Actualizar saldo del usuario
        usuario.saldo = saldo_final

        # ✅ Crear registro de recarga
        nueva_recarga = Recarga(
            usuario_id=usuario.id,
            saldo_anterior=saldo_anterior,
            monto_recargado=monto,
            saldo_final=saldo_final,
            referencia=referencia,
            fecha_recarga=datetime.now()
        )

        db.session.add(nueva_recarga)
        db.session.commit()

        return jsonify({
            "mensaje": f"Recarga exitosa para {usuario.nombre}",
            "saldo_anterior": saldo_anterior,
            "monto_recargado": monto,
            "nuevo_saldo": saldo_final,
            "referencia": referencia
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500


#Generar reporte recarga
@app.route('/usuarios/<int:usuario_id>/recargas', methods=['GET'])
def historial_recargas(usuario_id):
    try:
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({"message": "Usuario no encontrado"}), 404

        recargas = Recarga.query.filter_by(usuario_id=usuario.id).order_by(Recarga.fecha_recarga.desc()).all()

        historial = []
        for r in recargas:
            historial.append({
                "id": r.id,
                "saldo_anterior": r.saldo_anterior,
                "monto_recargado": r.monto_recargado,
                "saldo_final": r.saldo_final,
                "referencia": r.referencia,
                "fecha_recarga": r.fecha_recarga.strftime("%Y-%m-%d %H:%M:%S")
            })

        return jsonify({
            "usuario_id": usuario.id,
            "nombre": usuario.nombre,
            "recargas": historial
        }), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500
    
#consultar recargas todos los usuarios
@app.route('/recargas', methods=['GET'])
def obtener_recargas():
    try:
        recargas = Recarga.query.order_by(Recarga.fecha_recarga.desc()).all()
        formato = request.args.get("formato")

        if formato and formato.lower() == "excel":
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recargas"

            # Encabezados
            ws.append([
                "ID",
                "Usuario",
                "Número Identificación",
                "Saldo Anterior",
                "Monto Recargado",
                "Saldo Final",
                "Referencia",
                "Fecha Recarga"
            ])

            # Filas con datos
            for r in recargas:
                ws.append([
                    r.id,
                    r.usuario.nombre,
                    r.usuario.numero_identificacion,
                    r.saldo_anterior,
                    r.monto_recargado,
                    r.saldo_final,
                    r.referencia,
                    r.fecha_recarga.strftime("%Y-%m-%d %H:%M:%S")
                ])

            # Guardar en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                download_name="recargas.xlsx",
                as_attachment=True,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Si no piden Excel → devolver JSON
        recargas_list = []
        for r in recargas:
            recargas_list.append({
                "id": r.id,
                "usuario": r.usuario.nombre,
                "numero_identificacion": r.usuario.numero_identificacion,
                "saldo_anterior": r.saldo_anterior,
                "monto_recargado": r.monto_recargado,
                "saldo_final": r.saldo_final,
                "referencia": r.referencia,
                "fecha_recarga": r.fecha_recarga.strftime("%Y-%m-%d %H:%M:%S")
            })

        return jsonify(recargas_list), 200

    except Exception as e:
        return jsonify({"error": f"Error inesperado: {str(e)}"}), 500

# --------------------------
# ERRORES PERSONALIZADOS
# --------------------------
class VehiculoNoRegistradoError(Exception):
    pass # para cuando alguien intenta usar un vehículo que no existe en la BD

class SaldoInsuficienteError(Exception):
    pass # para cuando el saldo no alcanza para pagar la tarifa

class EspacioNoDisponibleError(Exception):
    pass # para cuando el parqueadero está lleno


# --------------------------
# MANEJO GLOBAL DE ERRORES
# --------------------------
def registrar_manejadores(app):

    # Error de integridad (duplicados, claves foráneas)
    @app.errorhandler(IntegrityError)
    def manejar_integrity_error(e):
        return jsonify({
            "error": "Violación de integridad en la base de datos",
            "detalle": str(e.orig)  # mensaje original de PostgreSQL
        }), 400

    # Error de datos inválidos (tipos, tamaños)
    @app.errorhandler(DataError)
    def manejar_data_error(e):
        return jsonify({
            "error": "Error en los datos enviados",
            "detalle": str(e.orig)
        }), 400

    # Error de conexión con la base de datos
    @app.errorhandler(OperationalError)
    def manejar_operational_error(e):
        return jsonify({
            "error": "Error de conexión con la base de datos",
            "detalle": str(e.orig)
        }), 500

    # Vehículo no encontrado
    @app.errorhandler(VehiculoNoRegistradoError)
    def manejar_vehiculo_no_registrado(e):
        return jsonify({
            "error": "Vehículo no registrado",
            "detalle": str(e)
        }), 404

    # Saldo insuficiente
    @app.errorhandler(SaldoInsuficienteError)
    def manejar_saldo_insuficiente(e):
        return jsonify({
            "error": "Saldo insuficiente",
            "detalle": str(e)
        }), 400

    # Espacio no disponible
    @app.errorhandler(EspacioNoDisponibleError)
    def manejar_espacio_no_disponible(e):
        return jsonify({
            "error": "No hay espacios disponibles",
            "detalle": str(e)
        }), 400

    # Error genérico (cualquier otro)
    @app.errorhandler(Exception)
    def manejar_error_general(e):
        return jsonify({
            "error": "Error inesperado en el servidor",
            "detalle": str(e)
        }), 500
    # Método HTTP no permitido
@app.errorhandler(MethodNotAllowed)
def manejar_method_not_allowed(e):
    return jsonify({
        "error": "Método HTTP no permitido",
        "detalle": f"Los métodos permitidos son: {', '.join(e.valid_methods) if e.valid_methods else 'desconocidos'}"
    }), 405
    

# ======================================================
# ENDPOINT RFID COMPLETO Y FUNCIONAL
# ======================================================

# ======================================================
# ENDPOINT RFID COMPLETO Y FUNCIONAL
# ======================================================

@app.route("/rfid", methods=["POST"])
def recibir_rfid():
    global ultimo_uid, ultimo_tipo, rfid_timestamp

    data = request.get_json()
    uid = data.get("uid")
    tipo = data.get("tipo")

    if not uid:
        return jsonify({"line1": "Error", "line2": "UID vacío"})

    # Guardar último UID leído
    ultimo_uid = uid
    ultimo_tipo = tipo
    rfid_timestamp = datetime.now()

    # ============================================================
    # MODO ASIGNACIÓN (solo mostrar el UID en pantalla)
    # ============================================================
    if tipo == "ASSIGN":
        return jsonify({
            "status": "OK",
            "line1": "RFID listo",
            "line2": uid
        })

    # ============================================================
    # Buscar vehículo asignado a ese RFID
    # ============================================================
    vehiculo = Vehiculo.query.filter_by(uid_rfid=uid).first()

    if not vehiculo:
        return jsonify({
            "status": "NO",
            "line1": "Acceso denegado",
            "line2": "RFID no registrado"
        })

    usuario = Usuario.query.get(vehiculo.usuario_id)

    # ============================================================
    # PROCESAR ENTRADA
    # ============================================================
    if tipo == "IN":
        registro_activo = Registro.query.filter_by(
            vehiculo_id=vehiculo.id, 
            hora_salida=None
        ).first()

        if registro_activo:
            return jsonify({
                "status": "NO",
                "line1": "Ya está adentro",
                "line2": "Use salida"
            })

        # Buscar espacio disponible
        espacio = Espacio.query.filter_by(
            tipo_vehiculo_id=vehiculo.tipo_vehiculo_id,
            estado=False
        ).first()

        if not espacio:
            return jsonify({
                "status": "NO",
                "line1": "Sin espacios",
                "line2": "Disponible"
            })

        # Ocupa el espacio
        espacio.estado = True
        espacio.vehiculo_id = vehiculo.id
        db.session.commit()

        # Crear registro
        registro = Registro(
            vehiculo_id=vehiculo.id,
            espacio_id=espacio.id,
            hora_ingreso=datetime.now()
        )
        db.session.add(registro)
        db.session.commit()

        return jsonify({
            "status": "OK_IN",
            "line1": "Bienvenido",
            "line2": f"{usuario.nombre[:16]} - Puesto {espacio.id}"
        })

    # ============================================================
    # PROCESAR SALIDA
    # ============================================================
    if tipo == "OUT":
        registro_activo = Registro.query.filter_by(
            vehiculo_id=vehiculo.id,
            hora_salida=None
        ).first()

        if not registro_activo:
            return jsonify({
                "status": "NO",
                "line1": "No está adentro",
                "line2": "Use entrada"
            })

        hora_salida = datetime.now()
        minutos = math.ceil((hora_salida - registro_activo.hora_ingreso).total_seconds() / 60)

        tarifa = Tarifa.query.filter_by(tipo_vehiculo_id=vehiculo.tipo_vehiculo_id).first()
        total_pago = minutos * tarifa.tarifa_hora / 60.0

        if usuario.saldo < total_pago:
            return jsonify({
                "status": "NO",
                "line1": "Saldo insuficiente",
                "line2": ""
            })

        # Cobro
        usuario.saldo -= total_pago
        registro_activo.hora_salida = hora_salida
        registro_activo.total_pago = total_pago
        registro_activo.tiempo_duracion = minutos

        # Liberar espacio
        espacio = Espacio.query.get(registro_activo.espacio_id)
        espacio.estado = False
        espacio.vehiculo_id = None

        db.session.commit()

        return jsonify({
            "status": "OK_OUT",
            "line1": "Hasta luego",
            "line2": usuario.nombre[:16]
        })

    # ============================================================
    # SI EL TIPO ES INVÁLIDO
    # ============================================================
    return jsonify({
        "status": "ERROR",
        "line1": "Tipo inválido",
        "line2": ""
    }), 400


# ======================================================
# EJECUCIÓN
# ======================================================
if __name__ == "__main__":
    with app.app_context():
        app.run(host="0.0.0.0", port=5000, debug=True)